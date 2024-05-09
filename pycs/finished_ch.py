import cv2
import requests
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
import re
import time

def extract_keyframe(video_url):
    response = requests.get(video_url, stream=True)
    if response.status_code != 200:
        print(f"请求URL {video_url} 失败，状态码：{response.status_code}")
        return None

    video_stream = BytesIO(response.content)
    cap = cv2.VideoCapture(video_stream)

    if not cap.isOpened():
        raise ValueError(f"无法从URL打开视频: {video_url}")

    ret, frame = cap.read()
    cap.release()
    return frame


def crop_and_save_image(frame, output_path):
    height, width, _ = frame.shape
    crop_size = int(min(height, width) * 0.25)
    x, y = (width - crop_size) // 2, (height - crop_size) // 2
    cropped_frame = frame[y:y + crop_size, x:x + crop_size]

    cv2.imwrite(output_path, cropped_frame)


def process_video_data(file_path="txt/finished_1.txt", output_file="xlsm/finished.xlsm"):
    valid_data = []
    invalid_lines = []

    with open(file_path, "r", encoding="utf-8") as f:
        total_lines = sum(1 for _ in f)
        f.seek(0)

        for i, line in enumerate(f, start=1):
            try:
                name, url, *_ = line.strip().split(",")
                valid_data.append({"name": name, "url": url})
            except Exception as e:
                invalid_lines.append((i, line.strip(), str(e)))

            print(f"正在处理第{i}/{total_lines}行，进度：{i/total_lines*100:.2f}%")
            time.sleep(0.1)  # 控制输出频率，避免刷屏

    print("\n无效行列表:")
    for line_no, line_content, error_msg in invalid_lines:
        print(f"第{line_no}行: {line_content} - 错误: {error_msg}\n")

    for item in valid_data:
        try:
            # 移除URL末尾的$及其右侧内容
            url = re.sub(r'\$.+$', '', item["url"])

            keyframe = extract_keyframe(url)
            name_base = item["name"]
            output_path = f"{name_base}_keyframe.jpg"
            crop_and_save_image(keyframe, output_path)
            item["cropped_keyframe"] = output_path
        except Exception as e:
            print(f"处理URL {item['url']} 时出错: {str(e)}")

    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    for item in valid_data:
        if "cropped_keyframe" in item:
            ws.append([item["name"], item["url"], item["cropped_keyframe"]])
        else:
            print(f"跳过缺少'cropped_keyframe'的项目: {item['name']}, {item['url']}")

    # 保存工作簿
    from datetime import datetime

    current_time = datetime.now()
    time_str = current_time.strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"workbook_{time_str}.xlsx"
    wb.save(output_file)


if __name__ == "__main__":
    process_video_data()
